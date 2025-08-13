"""
Professional Report Generator for S&P 500 Analysis Tool

This module provides comprehensive report generation capabilities including:
- PDF investment analysis reports
- Excel data export with formatting
- PowerPoint-ready chart exports
- Professional templates and branding

Author: Professional Investment Analysis Team
"""

import pandas as pd
import numpy as np
import io
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple
import base64

# PDF generation
from reportlab.lib.pagesizes import A4, letter
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm, inch
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.platypus import Image as RLImage
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont

# Excel generation
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# Chart generation
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots


class ReportTemplate:
    """Professional report template configuration."""
    
    def __init__(self):
        self.brand_color = "#0B3B5A"  # Deep navy blue
        self.accent_color = "#F59E0B"  # Amber
        self.success_color = "#16A34A"  # Green
        self.danger_color = "#DC2626"  # Red
        self.light_gray = "#F5F7FA"
        self.dark_gray = "#1F2937"
        
        self.fonts = {
            'title': ('Helvetica-Bold', 18),
            'heading': ('Helvetica-Bold', 14),
            'subheading': ('Helvetica-Bold', 12),
            'body': ('Helvetica', 10),
            'caption': ('Helvetica', 8)
        }
        
        # Ensure CJK font support
        self._setup_cjk_fonts()
    
    def _setup_cjk_fonts(self):
        """Setup Chinese font support."""
        try:
            pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
            self.cjk_font = "STSong-Light"
        except Exception:
            self.cjk_font = "Helvetica"


class ProfessionalPDFReport:
    """Professional PDF report generator."""
    
    def __init__(self, template: ReportTemplate = None):
        self.template = template or ReportTemplate()
        self.buffer = io.BytesIO()
        self.doc = None
        self.story = []
        
    def create_cover_page(self, title: str, subtitle: str, analysis_date: str, 
                         data_period: str, company_info: str = None) -> None:
        """Create professional cover page."""
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=24,
            spaceAfter=30,
            textColor=colors.HexColor(self.template.brand_color),
            alignment=TA_CENTER
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=16,
            spaceAfter=20,
            textColor=colors.HexColor(self.template.dark_gray),
            alignment=TA_CENTER
        )
        
        info_style = ParagraphStyle(
            'InfoStyle',
            parent=styles['Normal'],
            fontSize=12,
            spaceAfter=10,
            alignment=TA_CENTER
        )
        
        # Add cover content
        self.story.append(Spacer(1, 2*inch))
        self.story.append(Paragraph(title, title_style))
        self.story.append(Paragraph(subtitle, subtitle_style))
        self.story.append(Spacer(1, 1*inch))
        
        # Analysis information
        self.story.append(Paragraph(f"<b>分析日期 Analysis Date:</b> {analysis_date}", info_style))
        self.story.append(Paragraph(f"<b>数据期间 Data Period:</b> {data_period}", info_style))
        
        if company_info:
            self.story.append(Spacer(1, 0.5*inch))
            self.story.append(Paragraph(company_info, info_style))
        
        self.story.append(PageBreak())
    
    def add_executive_summary(self, summary_data: Dict[str, Any]) -> None:
        """Add executive summary section."""
        styles = getSampleStyleSheet()
        
        heading_style = ParagraphStyle(
            'Heading1',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=20,
            textColor=colors.HexColor(self.template.brand_color)
        )
        
        self.story.append(Paragraph("执行摘要 Executive Summary", heading_style))
        
        # Key metrics table
        key_metrics = summary_data.get('key_metrics', {})
        if key_metrics:
            data = [['指标 Metric', '数值 Value', '说明 Description']]
            
            for metric, info in key_metrics.items():
                data.append([
                    metric,
                    info.get('value', 'N/A'),
                    info.get('description', '')
                ])
            
            table = Table(data, colWidths=[2*inch, 1.5*inch, 3*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(self.template.brand_color)),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            self.story.append(table)
            self.story.append(Spacer(1, 20))
        
        # Summary text
        summary_text = summary_data.get('summary_text', '')
        if summary_text:
            self.story.append(Paragraph(summary_text, styles['Normal']))
            self.story.append(Spacer(1, 20))
    
    def add_risk_metrics_section(self, risk_data: Dict[str, Any]) -> None:
        """Add comprehensive risk metrics section."""
        styles = getSampleStyleSheet()
        
        heading_style = ParagraphStyle(
            'Heading1',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=20,
            textColor=colors.HexColor(self.template.brand_color)
        )
        
        self.story.append(Paragraph("风险指标分析 Risk Metrics Analysis", heading_style))
        
        # Overall risk metrics table
        overall_metrics = risk_data.get('overall_metrics', {})
        if overall_metrics:
            data = [['起始年份 Start Year', 'CAGR', '夏普比率 Sharpe', '最大回撤 Max DD', '波动率 Volatility']]
            
            for start_year, metrics in overall_metrics.items():
                data.append([
                    str(start_year),
                    f"{metrics.get('cagr', 0):.2%}",
                    f"{metrics.get('sharpe_ratio', 0):.3f}",
                    f"{metrics.get('max_drawdown', 0):.2%}",
                    f"{metrics.get('volatility', 0):.2%}"
                ])
            
            table = Table(data, colWidths=[1.2*inch, 1*inch, 1*inch, 1*inch, 1*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(self.template.brand_color)),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            self.story.append(table)
            self.story.append(Spacer(1, 20))
        
        # Risk interpretation
        risk_interpretation = risk_data.get('interpretation', '')
        if risk_interpretation:
            self.story.append(Paragraph(risk_interpretation, styles['Normal']))
            self.story.append(Spacer(1, 20))
    
    def add_chart_section(self, chart_data: bytes, title: str, description: str = "") -> None:
        """Add chart section with title and description."""
        styles = getSampleStyleSheet()
        
        subheading_style = ParagraphStyle(
            'Subheading',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=10,
            textColor=colors.HexColor(self.template.brand_color)
        )
        
        self.story.append(Paragraph(title, subheading_style))
        
        if description:
            self.story.append(Paragraph(description, styles['Normal']))
            self.story.append(Spacer(1, 10))
        
        # Add chart image
        chart_image = RLImage(io.BytesIO(chart_data), width=6*inch, height=4*inch)
        self.story.append(chart_image)
        self.story.append(Spacer(1, 20))
    
    def add_methodology_section(self) -> None:
        """Add methodology and disclaimers section."""
        styles = getSampleStyleSheet()
        
        heading_style = ParagraphStyle(
            'Heading1',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=20,
            textColor=colors.HexColor(self.template.brand_color)
        )
        
        self.story.append(Paragraph("方法论与免责声明 Methodology & Disclaimers", heading_style))
        
        methodology_text = """
        <b>计算方法 Calculation Methods:</b><br/>
        • 复合年增长率 (CAGR): 几何平均收益率<br/>
        • 夏普比率: (平均超额收益) / (收益波动率)<br/>
        • 最大回撤: 从峰值到谷底的最大损失<br/>
        • 波动率: 年化收益率标准差<br/><br/>
        
        <b>数据来源 Data Sources:</b><br/>
        • S&P 500 历史收益率数据来源于公开市场数据<br/>
        • 无风险利率基于美国国债收益率<br/><br/>
        
        <b>重要声明 Important Disclaimers:</b><br/>
        • 历史表现不代表未来结果<br/>
        • 投资涉及风险，可能导致本金损失<br/>
        • 本分析仅供参考，不构成投资建议<br/>
        • 请在做出投资决策前咨询专业财务顾问
        """
        
        self.story.append(Paragraph(methodology_text, styles['Normal']))
    
    def generate_pdf(self, filename: str = None) -> bytes:
        """Generate the complete PDF report."""
        if filename:
            self.doc = SimpleDocTemplate(filename, pagesize=A4)
        else:
            self.doc = SimpleDocTemplate(self.buffer, pagesize=A4)
        
        # Build the PDF
        self.doc.build(self.story)
        
        if filename:
            with open(filename, 'rb') as f:
                return f.read()
        else:
            return self.buffer.getvalue()


class ExcelReportGenerator:
    """Professional Excel report generator."""
    
    def __init__(self, template: ReportTemplate = None):
        self.template = template or ReportTemplate()
        self.workbook = openpyxl.Workbook()
        self.workbook.remove(self.workbook.active)  # Remove default sheet
    
    def create_summary_sheet(self, summary_data: Dict[str, Any]) -> None:
        """Create executive summary sheet."""
        ws = self.workbook.create_sheet("Executive Summary")
        
        # Title
        ws['A1'] = "S&P 500 Analysis - Executive Summary"
        ws['A1'].font = Font(size=16, bold=True, color="0B3B5A")
        ws.merge_cells('A1:E1')
        
        # Analysis date
        ws['A3'] = "Analysis Date:"
        ws['B3'] = datetime.now().strftime("%Y-%m-%d")
        
        # Key metrics
        row = 5
        ws[f'A{row}'] = "Key Metrics"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        
        headers = ['Metric', 'Value', 'Description']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row+1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="0B3B5A", end_color="0B3B5A", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Add metrics data
        metrics = summary_data.get('key_metrics', {})
        for i, (metric, info) in enumerate(metrics.items(), 1):
            ws.cell(row=row+1+i, column=1, value=metric)
            ws.cell(row=row+1+i, column=2, value=info.get('value', 'N/A'))
            ws.cell(row=row+1+i, column=3, value=info.get('description', ''))
        
        # Auto-adjust column widths
        for col_num in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_num)
            for row_num in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_num, column=col_num)
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def create_risk_metrics_sheet(self, risk_data: Dict[str, Any]) -> None:
        """Create detailed risk metrics sheet."""
        ws = self.workbook.create_sheet("Risk Metrics")
        
        # Title
        ws['A1'] = "Risk Metrics Analysis"
        ws['A1'].font = Font(size=16, bold=True, color="0B3B5A")
        
        # Headers
        headers = ['Start Year', 'CAGR', 'Sharpe Ratio', 'Sortino Ratio', 'Calmar Ratio', 
                  'Volatility', 'Max Drawdown', 'VaR (95%)', 'CVaR (95%)']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="0B3B5A", end_color="0B3B5A", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Add data
        overall_metrics = risk_data.get('overall_metrics', {})
        for row, (start_year, metrics) in enumerate(overall_metrics.items(), 4):
            ws.cell(row=row, column=1, value=start_year)
            ws.cell(row=row, column=2, value=metrics.get('cagr', 0))
            ws.cell(row=row, column=3, value=metrics.get('sharpe_ratio', 0))
            ws.cell(row=row, column=4, value=metrics.get('sortino_ratio', 0))
            ws.cell(row=row, column=5, value=metrics.get('calmar_ratio', 0))
            ws.cell(row=row, column=6, value=metrics.get('volatility', 0))
            ws.cell(row=row, column=7, value=metrics.get('max_drawdown', 0))
            ws.cell(row=row, column=8, value=metrics.get('var_95', 0))
            ws.cell(row=row, column=9, value=metrics.get('cvar_95', 0))
        
        # Format percentage columns
        percentage_cols = [2, 6, 7, 8, 9]  # CAGR, Volatility, Max DD, VaR, CVaR
        for col in percentage_cols:
            for row in range(4, 4 + len(overall_metrics)):
                ws.cell(row=row, column=col).number_format = '0.00%'
        
        # Format ratio columns
        ratio_cols = [3, 4, 5]  # Sharpe, Sortino, Calmar
        for col in ratio_cols:
            for row in range(4, 4 + len(overall_metrics)):
                ws.cell(row=row, column=col).number_format = '0.000'
    
    def generate_excel(self, filename: str = None) -> bytes:
        """Generate the complete Excel report."""
        if filename:
            self.workbook.save(filename)
            with open(filename, 'rb') as f:
                return f.read()
        else:
            buffer = io.BytesIO()
            self.workbook.save(buffer)
            return buffer.getvalue()


class ChartExporter:
    """High-quality chart export functionality."""
    
    @staticmethod
    def export_risk_metrics_chart(risk_data: Dict[str, Any], format: str = 'png') -> bytes:
        """Export risk metrics comparison chart."""
        overall_metrics = risk_data.get('overall_metrics', {})
        
        if not overall_metrics:
            return b''
        
        # Create subplots
        fig = make_subplots(
            rows=2, cols=2,
            subplot_titles=["CAGR vs Volatility", "Sharpe Ratio", "Maximum Drawdown", "Risk-Return Profile"],
            specs=[[{"secondary_y": False}, {"secondary_y": False}],
                   [{"secondary_y": False}, {"secondary_y": False}]]
        )
        
        start_years = list(overall_metrics.keys())
        cagrs = [overall_metrics[sy].get('cagr', 0) for sy in start_years]
        volatilities = [overall_metrics[sy].get('volatility', 0) for sy in start_years]
        sharpe_ratios = [overall_metrics[sy].get('sharpe_ratio', 0) for sy in start_years]
        max_drawdowns = [overall_metrics[sy].get('max_drawdown', 0) for sy in start_years]
        
        # CAGR vs Volatility scatter
        fig.add_trace(
            go.Scatter(x=volatilities, y=cagrs, mode='markers+text', 
                      text=[str(sy) for sy in start_years], textposition="top center",
                      name="Risk-Return", marker=dict(size=10, color="#0B3B5A")),
            row=1, col=1
        )
        
        # Sharpe Ratio bar chart
        fig.add_trace(
            go.Bar(x=[str(sy) for sy in start_years], y=sharpe_ratios, 
                   name="Sharpe Ratio", marker_color="#F59E0B"),
            row=1, col=2
        )
        
        # Maximum Drawdown bar chart
        fig.add_trace(
            go.Bar(x=[str(sy) for sy in start_years], y=max_drawdowns, 
                   name="Max Drawdown", marker_color="#DC2626"),
            row=2, col=1
        )
        
        # Risk-Return profile
        fig.add_trace(
            go.Scatter(x=volatilities, y=sharpe_ratios, mode='markers+text',
                      text=[str(sy) for sy in start_years], textposition="top center",
                      name="Volatility vs Sharpe", marker=dict(size=10, color="#16A34A")),
            row=2, col=2
        )
        
        # Update layout
        fig.update_layout(
            height=800,
            showlegend=False,
            title_text="Risk Metrics Analysis",
            title_x=0.5
        )
        
        # Export as bytes
        try:
            if format.lower() == 'svg':
                return fig.to_image(format="svg")
            else:
                return fig.to_image(format="png", width=1200, height=800, scale=2)
        except Exception as e:
            # Fallback: create a simple chart if plotly export fails
            print(f"Warning: Chart export failed: {e}")
            # Return empty bytes for now
            return b''


def generate_comprehensive_report(
    analysis_results: Dict[str, Any],
    config: Dict[str, Any],
    report_type: str = 'pdf'
) -> bytes:
    """Generate comprehensive professional report."""
    
    if report_type.lower() == 'pdf':
        # Generate PDF report
        template = ReportTemplate()
        pdf_report = ProfessionalPDFReport(template)
        
        # Cover page
        analysis_date = datetime.now().strftime("%Y-%m-%d")
        data_period = f"{min(config.get('start_years', [1926]))}-{datetime.now().year}"
        
        pdf_report.create_cover_page(
            title="S&P 500 投资分析报告\nS&P 500 Investment Analysis Report",
            subtitle="滚动收益率与风险指标分析\nRolling Returns & Risk Metrics Analysis",
            analysis_date=analysis_date,
            data_period=data_period,
            company_info="Professional Investment Analysis Platform"
        )
        
        # Executive summary
        summary_data = {
            'key_metrics': {
                '分析期间 Analysis Period': {'value': data_period, 'description': '数据覆盖期间'},
                '起始年份数量 Start Years': {'value': len(config.get('start_years', [])), 'description': '分析的基准年份数量'},
                '时间窗口 Time Windows': {'value': len(config.get('windows', [])), 'description': '滚动分析窗口数量'}
            },
            'summary_text': '本报告基于历史S&P 500数据，提供全面的风险收益分析，包括滚动收益率、风险指标和收敛性分析。'
        }
        pdf_report.add_executive_summary(summary_data)
        
        # Risk metrics section
        if 'risk_metrics' in analysis_results:
            risk_data = {
                'overall_metrics': {},
                'interpretation': '风险指标分析显示了不同时期的风险收益特征，为投资决策提供重要参考。'
            }
            
            # Process risk metrics data
            for start_year, year_data in analysis_results['risk_metrics'].items():
                if 'overall' in year_data and year_data['overall']:
                    risk_data['overall_metrics'][start_year] = year_data['overall']
            
            pdf_report.add_risk_metrics_section(risk_data)
        
        # Methodology section
        pdf_report.add_methodology_section()
        
        return pdf_report.generate_pdf()
    
    elif report_type.lower() == 'excel':
        # Generate Excel report
        template = ReportTemplate()
        excel_report = ExcelReportGenerator(template)
        
        # Summary sheet
        summary_data = {
            'key_metrics': {
                'Analysis Period': {'value': f"{min(config.get('start_years', [1926]))}-{datetime.now().year}", 'description': 'Data coverage period'},
                'Start Years Count': {'value': len(config.get('start_years', [])), 'description': 'Number of baseline years analyzed'},
                'Time Windows': {'value': len(config.get('windows', [])), 'description': 'Number of rolling analysis windows'}
            }
        }
        excel_report.create_summary_sheet(summary_data)
        
        # Risk metrics sheet
        if 'risk_metrics' in analysis_results:
            risk_data = {'overall_metrics': {}}
            for start_year, year_data in analysis_results['risk_metrics'].items():
                if 'overall' in year_data and year_data['overall']:
                    risk_data['overall_metrics'][start_year] = year_data['overall']
            
            excel_report.create_risk_metrics_sheet(risk_data)
        
        return excel_report.generate_excel()
    
    else:
        raise ValueError(f"Unsupported report type: {report_type}")
